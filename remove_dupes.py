from openpyxl import load_workbook
import time

# Define the path to your input Excel file
input_file = 'Duplicate_Records.xlsx'
# Define the path to your output Excel file
output_file = 'Cleaned_Records.xlsx'

# Load the workbook
wb = load_workbook(filename=input_file) 
# Select the active worksheet
ws = wb.active

# Get the column headers from the first row
column_headers = [cell.value for cell in ws[1]]

# Define the columns to avoid
values_to_remove = ['KEEP OR DELETE','Version','TransactionID']

# Define the columns to consider for uniqueness
unique_columns = [x for x in column_headers if x not in values_to_remove]

# Dictionary to store the maximum txn_id for each unique combination of other fields
max_txn_ids = {}

# Find the index of the column headers specified in unique_columns
column_indices = [column_headers.index(col) for col in unique_columns]

# Iterate through the rows to find the maximum txn_id for each unique combination of other fields
for row in ws.iter_rows(min_row=2, values_only=True):
    key = tuple(row[i] for i in column_indices)

    txn_id = int(row[0]) #'txn_id' is in the first column (A)
    
    if key in max_txn_ids:
        max_txn_ids[key] = max(max_txn_ids[key], txn_id)
    else:
        max_txn_ids[key] = txn_id

# Iterate through the rows again to filter and remove duplicates
rows_to_delete = set()
for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    key = tuple(row[i] for i in column_indices)
    txn_id = int(row[0])  #'txn_id' is in the first column (A)
    if txn_id != max_txn_ids[key]:
        rows_to_delete.add(idx)

# Delete rows with duplicate values
for row_index in sorted(rows_to_delete, reverse=True):
    ws.delete_rows(row_index)

# Save the modified workbook
wb.save(output_file)
